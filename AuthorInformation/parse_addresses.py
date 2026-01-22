"""

"""

import os
import sys
import argparse
from author import Author, Address
import operator
from collections import OrderedDict
from openpyxl import load_workbook


def parse_row(row, verbose=False):
    """
    parse a row from a tsv file or an excel file
    """

    if len(row) < 15:
        sys.stderr.write("ERROR: Malformed: {}\t{}\n".format(len(row), row))
        return None

    if not row[2]:
        return None

    auth = Author(row[2], verbose=verbose)
    if verbose:
        print(f"Auth: {auth}", file=sys.stderr)
    if len(row) > 24 and row[24]:
        auth.funding = row[2] + " " + row[24].replace('"', '')

    try:
        if row[1]:
            auth.orcid = row[1]
        if row[3]:
            auth.lastname = row[3]
            auth.lastnamelower = row[3].lower()
        if row[4]:
            auth.firstname = row[4]
            auth.firstnamelower = row[4].lower()
        if row[5]:
            auth.middleinitial = row[5]
        if row[6]:
            auth.email = row[6].replace(' ', '')
        if row[14]:
            auth.order = int(row[14])
        if row[15]:
            auth.contribution = row[15].replace('"', '')
    except Exception as err:
        print(f"Error parsing author name {row[2]}: {err}\n", file=sys.stderr)
        return None

    if verbose:
        print(f"Checking primary address for {auth}", file=sys.stderr)

    primary = Address(verbose=verbose)
    try:
        if row[7]:
            primary.department = row[7].replace('"', '')
        if row[8]:
            primary.institution = row[8].replace('"', '')
        if row[9]:
            primary.street = row[9].replace('"', '')
        if row[10]:
            primary.city = row[10].replace('"', '')
        if row[11]:
            primary.state = row[11].replace('"', '')
        if row[12]:
            primary.zip = str(row[12]).replace('"', '')
        if row[13]:
            primary.country = row[13].replace('"', '')
    except Exception as err:
        print(f"Error parsing primary address for {row[2]}: {err}\n", file=sys.stderr)
        return None

    auth.primaryaddress = primary
    if verbose:
        print(f"Checking secondary address for {auth}", file=sys.stderr)

    secondary = Address(verbose=verbose)
    try:
        if len(row) > 17 and row[17]:
            secondary.department = row[17].replace('"', '')
        if len(row) > 18 and row[18]:
            secondary.institution = row[18].replace('"', '')
        if len(row) > 19 and row[19]:
            secondary.street = row[19].replace('"', '')
        if len(row) > 20 and row[20]:
            secondary.city = row[20].replace('"', '')
        if len(row) > 21 and row[21]:
            secondary.state = row[21].replace('"', '')
        if len(row) > 22 and row[22]:
            secondary.zip = str(row[22]).replace('"', '')
        if len(row) > 23 and row[23]:
            secondary.country = row[23].replace('"', '')

        if secondary.is_valid():
            auth.secondaryaddress = secondary
    except Exception as err:
        print(f"Error parsing secondary address for {row[2]}: {err}\n", file=sys.stderr)
        return None

    return auth


def parse_file(filename, verbose=False):
    """
    parse a file and create a set of authors
    :param filename: file to parse
    :return: set of author elements
    """

    authors = set()
    abbrevs = set()
    firstline = True # ignore the first line

    if '.xls' in filename:
        try:
            workbook = load_workbook(filename=filename, data_only=True)
            sheet = workbook["authors"]
        except Exception as err:
            print(f"Error parsing {filename} as an excel file and extracting sheet 'authors': {err}", file=sys.stderr)
            return None
        for row in sheet.iter_rows(values_only=True):
            if row[1] == 'ORCID' or 'acronym' in row[2]:
                continue
            auth = parse_row(row, verbose=verbose)
            if auth in authors:
                print(f"Duplicate author {auth} in {filename}", file=sys.stderr)
                continue
            authors.add(auth)
    else:
        with open(filename, 'r', encoding='cp1252') as f:
            for l in f:
                if firstline:
                    firstline = False
                    continue
                l = l.rstrip()
                if not l:
                    continue
                row = l.split("\t")
                auth = parse_row(row, verbose=verbose)
                if auth in authors:
                    print(f"Duplicate author {auth} in {filename}", file=sys.stderr)
                    continue
                authors.add(auth)
    return authors


def test_validity(authors):
    """
    Test whether the author information is valid
    :param authors: the set of authors
    :return: nothing
    """

    abbs = set()

    for a in authors:
        if a.abbreviation.lower() in abbs:
            sys.stderr.write("FATAL: {} is a duplicate abbrevation\n".format(a.abbreviation))
        abbs.add(a.abbreviation.lower())
        if not a.is_valid():
            a.verbose = True
        print("{}\t{}\t{}".format(a.abbreviation, a.get_name(), a.is_valid()))

def check_spellings(authors):
    """
    Check for potential misspellings based on institutions, departments, and zip codes
    :param authors: the list of authors
    :return:
    """

    # find the set of different addresses
    addresses = {}
    allinst = {}
    allzip = {}
    for a in sorted(authors, key=operator.attrgetter('order', 'lastnamelower', 'firstnamelower')):
        pa = a.primaryaddress.get_address()
        if pa not in addresses:
            addresses[pa] = {}
            addresses[pa]['institution'] = a.primaryaddress.institution
            if a.primaryaddress.institution not in allinst:
                allinst[a.primaryaddress.institution] = 1
            else:
                allinst[a.primaryaddress.institution] += 1
            addresses[pa]['zip'] = a.primaryaddress.zip
            if a.primaryaddress.zip not in allzip:
                allzip[a.primaryaddress.zip] = 1
            else:
                allzip[a.primaryaddress.zip] += 1

        if a.secondaryaddress.is_valid():
            sa = a.secondaryaddress.get_address()
            if sa not in addresses:
                addresses[sa] = {}
                addresses[sa]['institution'] = a.secondaryaddress.institution
                if a.secondaryaddress.institution not in allinst:
                    allinst[a.secondaryaddress.institution] = 1
                else:
                    allinst[a.secondaryaddress.institution] += 1
                addresses[sa]['zip'] = a.secondaryaddress.zip
                if a.secondaryaddress.zip not in allzip:
                    allzip[a.secondaryaddress.zip] = 1
                else:
                    allzip[a.secondaryaddress.zip] += 1

    sys.stderr.write("Duplicates by institution\n")
    for i in allinst:
        if not i:
            continue
        if allinst[i] < 2:
            continue
        sys.stderr.write("\n")
        for a in addresses:
            if addresses[a]['institution'] == i:
                sys.stderr.write("{}\n".format(a))

    sys.stderr.write("\n\n\nDuplicates by zip\n")
    for z in allzip:
        if not z:
            continue
        if allzip[z] < 2:
            continue
        sys.stderr.write("\n")
        for a in addresses:
            if addresses[a]['zip'] == z:
                sys.stderr.write("{}\n".format(a))


def print_author_list(authors):
    """
    Print the list of all authors.
    :param authors: the set of authors
    :return:
    """

    # the list of addresses as we add them. This becomes the order
    addresses = []
    a: Author
    sys.stdout.write("<p>")
    for a in sorted(authors, key=operator.attrgetter('order', 'lastnamelower', 'firstnamelower')):
        output = a.get_name()
        pa = a.primaryaddress.get_address()
        if pa not in addresses:
            addresses.append(pa)
        addidx = addresses.index(pa) + 1
        if a.secondaryaddress.is_valid():
            sa = a.secondaryaddress.get_address()
            if sa not in addresses:
                addresses.append(sa)
            oidx = addresses.index(sa) + 1
            addidx = "{},{}".format(addidx, oidx)
        output += "<sup>{}</sup>, ".format(addidx)
        sys.stdout.write(output)

    sys.stdout.write("</p>\n\n\n")
    for i, j in enumerate(addresses):
        print("<sup>{}</sup>{}<br>".format(i+1,j))

def print_author_contributions(authors):
    """
    Print the author contribution list
    :param authors:
    :return:
    """

    contribs = OrderedDict()


    a: Author
    for a in sorted(authors, key=operator.attrgetter('order', 'lastnamelower', 'firstnamelower')):
        c = a.contribution
        thisc = ''.join([c[0].lower(), c[1:]]) # convert the first letter to lower case as it will be in a sentence
        if thisc not in contribs:
            contribs[thisc] = []
        contribs[thisc].append(a.abbreviation)

    sys.stdout.write("<p> &nbsp; </p><h1>Author Contributions</h1><p> &nbsp; </p>\n")
    for c in contribs:
        output = ", ".join(map(str, sorted(contribs[c])))
        output += f" {c}"
        sys.stdout.write(f"{output}. ")
    sys.stdout.write("</p>\n\n\n")

def print_funding(authors):
    """
    Print the author funding list
    :param authors:
    :return:
    """

    funding = OrderedDict()

    a: Author
    for a in sorted(authors, key=operator.attrgetter('order', 'lastnamelower', 'firstnamelower')):
        f = a.funding
        if not f:
            continue
        # thisf = ''.join([f[0].lower(), f[1:]]) # convert the first letter to lower case as it will be in a sentence
        thisf = f
        if thisf not in funding:
            funding[thisf] = []
        funding[thisf].append(a.abbreviation)

    sys.stdout.write("<p> &nbsp; </p><h1>Funding Acknowledgements</h1><p> &nbsp; </p>\n")
    for f in funding:
        output = ", ".join(map(str, sorted(funding[f])))
        output += f" acknowledges funding from {f}"
        sys.stdout.write(f"{output}. ")
    sys.stdout.write("</p>\n\n\n")

def print_orcids(authors):
    """
    Print the ORCID list
    """

    print("\n\n<h1>ORCID Information</h1>\n<ul>")
    for a in sorted(authors, key=operator.attrgetter('order', 'lastnamelower', 'firstnamelower')):
        print(f"<li>{a.abbreviation}: {a.orcid}</li>")
    print("</ul>")

def nature_form(authors):
    """
    Print a text list of authors to be cut and pasted for nature

    First Name
    Middle Initial
    Last Name
    email
    Institution
    City
    Country

    :param authors: the set of authors
    :return:
    """
    counter=0
    for a in sorted(authors, key=operator.attrgetter('order', 'lastnamelower', 'firstnamelower')):
        if None == a.middleinitial:
            a.middleinitial = ""
        print(counter)
        print("\n".join(map(str, [a.firstname, a.middleinitial, a.lastname, a.email,
                         a.primaryaddress.institution, a.primaryaddress.city,
                         a.primaryaddress.country])))
        print("\n\n")
        counter+=1

def science_list(authors):
    """
    Print a list of the authors for science. This is easy, first, last, email
    :param authors: the set of authors
    :return:
    """
    for a in sorted(authors, key=operator.attrgetter('order', 'lastnamelower', 'firstnamelower')):
        print("\t".join(map(str, [a.firstname, a.lastname, a.email])))

def comma_list(authors):
    """
    Print a list of the authors with first MI last,
    :param authors: the set of authors
    :return:
    """
    for a in sorted(authors, key=operator.attrgetter('order', 'lastnamelower', 'firstnamelower')):
        if a.middleinitial:
            print(" ".join(map(str, [a.firstname, a.middleinitial + ".", a.lastname])), end=", ")
        else:
            print(" ".join(map(str, [a.firstname, a.lastname])), end=", ")

def biorxiv(authors):
    """
    print a tsv file that you can upload to biorxiv
    Fields are Email	Institution	First Name	Middle Name(s)/Initial(s)	Last Name	Suffix	Corresponding Author	Home Page URL	Collaborative Group/Consortium	ORCiD

    :param authors: the set of authors
    :return:
    """

    print("Email\tInstitution\tFirst Name\tMiddle Name(s)/Initial(s)\tLast Name\tSuffix\tCorresponding Author\tHome Page URL\tCollaborative Group/Consortium\tORCiD")

    for a in sorted(authors, key=operator.attrgetter('order', 'lastnamelower', 'firstnamelower')):
        data= [a.email]
        inst = a.primaryaddress.get_address()
        secinst = a.secondaryaddress.get_address()
        if secinst:
            inst = f"{inst}; {secinst}"
        data.append(inst)
        data+= [a.firstname, a.middleinitial, a.lastname, "", "", "", "", a.orcid]
        print("\t".join(map(str, data)))



def email_list(authors):
    """
    Print a list suitable for emailing all colleagues.
    :param authors: the set of authors
    :return:
    """
    for a in sorted(authors, key=operator.attrgetter('order', 'lastnamelower', 'firstnamelower')):
        sys.stdout.write("{} {} <{}>, ".format(a.firstname, a.lastname, a.email))
    print()


def orcid_list(authors):
    """
    Print the author list in the correct order, but print their orcid
    :param authors: the set of authors
    :return:
    """

    for a in sorted(authors, key=operator.attrgetter('order', 'lastnamelower', 'firstnamelower')):
        orcid = a.orcid
        nm = a.get_name()
        if orcid:
            print(f"{orcid}\t{nm}")
        else:
            sys.stderr.write("No ORCID for {}\n".format(nm))
            print(nm)

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description="Parse the author information from our google doc. We can read either an excel or tsv file")
    parser.add_argument('-f', help='Author information', required=True)
    parser.add_argument('-t', help="test validity of the author information", action='store_true')
    parser.add_argument('-d', help='check for duplicate entries', action='store_true')
    parser.add_argument('-o', help='print the author list as ORCids in the correct order', action='store_true')
    parser.add_argument('-n', help='print the author list suitable for cutting and pasting to nature', action='store_true')
    parser.add_argument('-s', help='print the author list to add to the science bulk upload', action='store_true')
    parser.add_argument('-b', help='print a tsv file to upload to bioRxiv', action='store_true')
    parser.add_argument('-c', help='print the author list comma separated', action='store_true')
    parser.add_argument('-e', help='print the author list to use sending emails', action='store_true')
    parser.add_argument('-v', help='verbose output', action="store_true")
    args = parser.parse_args()

    authors = parse_file(args.f, args.v)

    if args.d:
        check_spellings(authors)
        sys.exit(0)

    if args.t:
        test_validity(authors)
        sys.exit(0)

    if args.n:
        nature_form(authors)
        sys.exit(0)

    if args.s:
        science_list(authors)
        sys.exit(0)

    if args.c:
        comma_list(authors)
        print()
        sys.exit(0)

    if args.b:
        biorxiv(authors)
        print()
        sys.exit(0)

    if args.e:
        email_list(authors)
        sys.exit(0)

    if args.o:
        orcid_list(authors)
        sys.exit(0)

    print_author_list(authors)
    print_orcids(authors)
    print_author_contributions(authors)
    print_funding(authors)



